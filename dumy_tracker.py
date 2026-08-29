"""
Daily email tracker automation.

Pulls sent Gmail messages within a time window and writes one row per email
to dumy_tracker.xlsx, matching the manual tracker's columns:
Date | Employee Name | Task Type | Task/Ticket Description or Reference No. | Start Time | End Time

Edit START_TIME / END_TIME below, then run:
    python dumy_tracker.py
"""

import datetime as dt
import os
import re

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from openpyxl import Workbook, load_workbook

# Edit these two values before each run, format: "YYYY-MM-DD HH:MM"
START_TIME = "2026-08-28 09:00"
END_TIME = "2026-08-28 19:00"

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
CREDENTIALS_PATH = r"C:\Users\isaac\credentials.json"
TOKEN_PATH = r"C:\Users\isaac\token.json"
TRACKER_PATH = r"C:\Users\isaac\dumy_tracker.xlsx"

HEADERS = [
    "Date",
    "Employee Name",
    "Task Type",
    "Task/Ticket Description or Reference No.",
    "Start Time",
    "End Time",
]


def get_gmail_service():
    creds = None
    if os.path.exists(TOKEN_PATH):
        creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_PATH, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_PATH, "w") as token_file:
            token_file.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)


def get_display_name(service):
    profile = service.users().getProfile(userId="me").execute()
    email_address = profile["emailAddress"]
    # Try to derive a friendlier name from a recent sent message header.
    results = service.users().messages().list(userId="me", q="in:sent", maxResults=1).execute()
    messages = results.get("messages", [])
    if messages:
        msg = service.users().messages().get(
            userId="me", id=messages[0]["id"], format="metadata", metadataHeaders=["From"]
        ).execute()
        for header in msg["payload"].get("headers", []):
            if header["name"] == "From":
                match = re.match(r'^"?([^"<]+)"?\s*<', header["value"])
                if match:
                    return match.group(1).strip()
    return email_address


def extract_header(headers, name):
    for header in headers:
        if header["name"].lower() == name.lower():
            return header["value"]
    return ""


def parse_recipient_name(to_header):
    match = re.match(r'^"?([^"<]+)"?\s*<', to_header)
    if match:
        return match.group(1).strip()
    match = re.match(r"^([^@<>]+)@", to_header)
    if match:
        return match.group(1).strip()
    return to_header.strip()


def parse_configured_datetime(value, label):
    try:
        return dt.datetime.strptime(value, "%Y-%m-%d %H:%M")
    except ValueError:
        raise SystemExit(f"{label} must look like 'YYYY-MM-DD HH:MM' (got: {value!r})")


def get_time_range():
    start_dt = parse_configured_datetime(START_TIME, "START_TIME")
    end_dt = parse_configured_datetime(END_TIME, "END_TIME")
    if end_dt <= start_dt:
        raise SystemExit("END_TIME must be after START_TIME")
    return start_dt, end_dt


def get_sent_emails_in_range(service, start_dt, end_dt):
    # Gmail search only filters by whole day, so query the days the window
    # spans, then filter to the exact start/end timestamps in Python.
    query_after = start_dt.date().strftime("%Y/%m/%d")
    query_before = (end_dt.date() + dt.timedelta(days=1)).strftime("%Y/%m/%d")
    query = f"in:sent after:{query_after} before:{query_before}"

    emails = []
    request = service.users().messages().list(userId="me", q=query)
    while request is not None:
        response = request.execute()
        for msg_ref in response.get("messages", []):
            msg = service.users().messages().get(
                userId="me",
                id=msg_ref["id"],
                format="metadata",
                metadataHeaders=["Subject", "To", "Date"],
            ).execute()
            sent_dt = dt.datetime.fromtimestamp(int(msg["internalDate"]) / 1000)
            if not (start_dt <= sent_dt <= end_dt):
                continue
            headers = msg["payload"].get("headers", [])
            subject = extract_header(headers, "Subject") or "(no subject)"
            to_header = extract_header(headers, "To") or "(unknown recipient)"
            recipient = parse_recipient_name(to_header)
            emails.append(
                {
                    "date": sent_dt.date(),
                    "time": sent_dt,
                    "description": f"{subject} \u2014 to {recipient}",
                }
            )
        request = service.users().messages().list_next(previous_request=request, previous_response=response)

    emails.sort(key=lambda e: e["time"])
    return emails


def load_or_create_tracker():
    if os.path.exists(TRACKER_PATH):
        wb = load_workbook(TRACKER_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tracker"
        ws.append(HEADERS)
    return wb, ws


def clear_data_rows(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def append_rows(ws, employee_name, emails):
    added = 0
    for email in emails:
        date_str = email["time"].strftime("%m/%d/%Y")
        time_str = email["time"].strftime("%m/%d/%Y %H:%M:%S")
        ws.append([date_str, employee_name, "Email Ticket", email["description"], time_str, time_str])
        added += 1
    return added


def main():
    start_dt, end_dt = get_time_range()

    service = get_gmail_service()
    employee_name = get_display_name(service)
    emails = get_sent_emails_in_range(service, start_dt, end_dt)

    if not emails:
        print(f"No sent emails found between {start_dt} and {end_dt}.")
        return

    wb, ws = load_or_create_tracker()
    clear_data_rows(ws)
    added = append_rows(ws, employee_name, emails)
    wb.save(TRACKER_PATH)
    print(f"Added {added} email entries ({start_dt} to {end_dt}) to {TRACKER_PATH}")


if __name__ == "__main__":
    main()
